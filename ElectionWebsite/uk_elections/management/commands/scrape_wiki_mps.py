"""
Scrape Wikipedia "List of MPs elected" pages and import elected candidates
into CandidateResult for pre-1832 general elections.

Usage:
    # Dry run for one year — shows matched/unmatched names and seat mismatches
    python manage.py scrape_wiki_mps --year 1826 --dry-run

    # Import for one year
    python manage.py scrape_wiki_mps --year 1826

    # Import all pre-1832 GEs
    python manage.py scrape_wiki_mps --all

    # Generate a review spreadsheet without touching the DB
    python manage.py scrape_wiki_mps --all --dry-run --spreadsheet review.xlsx

    # Import and also save spreadsheet
    python manage.py scrape_wiki_mps --all --spreadsheet review.xlsx

    # Overwrite existing CandidateResult rows (re-import)
    python manage.py scrape_wiki_mps --year 1826 --overwrite

    # Also try to fill in missing parties from individual constituency pages
    # (slow — one HTTP request per constituency with a missing party)
    python manage.py scrape_wiki_mps --year 1826 --fetch-parties

Add constituency name corrections to CONSTITUENCY_ALIASES (Wikipedia name -> DB name).
Add party name corrections to PARTY_ALIASES (Wikipedia name lowercased -> DB name).
"""

import re
import time
from collections import Counter

import requests
from bs4 import BeautifulSoup

from django.core.management.base import BaseCommand
from django.db.models import Q

from uk_elections.models import Election, Constituency, Party, CandidateResult

# ── Name corrections ──────────────────────────────────────────────────────────
CONSTITUENCY_ALIASES = {
    # 'Antrim':           'County Antrim',
    # 'Armagh':           'County Armagh',
    # 'St Germans':       'St. Germans',
}

PARTY_ALIASES = {
    # 'court whig':       'Whig',
    # 'opp. whig':        'Whig',
}

# ── Wikipedia URL patterns ────────────────────────────────────────────────────
URL_TEMPLATES = [
    "https://en.wikipedia.org/wiki/List_of_MPs_elected_in_the_{year}_United_Kingdom_general_election",
    "https://en.wikipedia.org/wiki/List_of_MPs_elected_in_the_United_Kingdom_general_election,_{year}",
    "https://en.wikipedia.org/wiki/List_of_MPs_elected_in_the_{year}_British_general_election",
    "https://en.wikipedia.org/wiki/List_of_MPs_elected_in_the_British_general_election,_{year}",
    "https://en.wikipedia.org/wiki/List_of_MPs_elected_in_the_{year}_English_general_election",
    "https://en.wikipedia.org/wiki/List_of_MPs_elected_in_the_English_general_election,_{year}",
]

HEADERS = {'User-Agent': 'ElectionWebsite-research-bot/1.0 (educational project)'}

_SEAT_SUFFIX = re.compile(
    r'\s*\(\s*(?:\d+\s+members?|two\s+members?|seat\s+\d+/\d+)\s*\)',
    re.IGNORECASE,
)
_FOOTNOTE = re.compile(r'\[.*?\]')


def _clean(text):
    text = _FOOTNOTE.sub('', text)
    text = _SEAT_SUFFIX.sub('', text)
    return text.strip()


# ── HTTP helpers ──────────────────────────────────────────────────────────────

def find_wiki_page(year):
    for template in URL_TEMPLATES:
        url = template.format(year=year)
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        except requests.RequestException:
            continue
        if resp.status_code == 200:
            return resp.url, BeautifulSoup(resp.text, 'html.parser')
    return None, None


_const_party_cache: dict[str, dict[str, str]] = {}


def _fetch_const_page_parties(const_name: str) -> dict[str, str]:
    wiki_title = const_name.replace(' ', '_') + '_(UK_Parliament_constituency)'
    url = f'https://en.wikipedia.org/wiki/{wiki_title}'
    try:
        resp = requests.get(url, headers=HEADERS, timeout=8, allow_redirects=True)
    except requests.RequestException:
        return {}
    if resp.status_code != 200:
        return {}

    soup = BeautifulSoup(resp.text, 'html.parser')
    result: dict[str, str] = {}

    for table in soup.find_all('table', class_='wikitable'):
        headers = [th.get_text(strip=True).lower() for th in table.find_all('th')]
        if 'party' not in ' '.join(headers):
            continue
        party_indices = [i for i, h in enumerate(headers) if 'party' in h]
        member_indices = [i for i, h in enumerate(headers)
                          if 'member' in h or h in ('mp', 'name')]
        for tr in table.find_all('tr'):
            cells = tr.find_all('td')
            if not cells:
                continue
            cell_texts = [_FOOTNOTE.sub('', c.get_text(strip=True)) for c in cells]
            for mi, pi in zip(member_indices, party_indices):
                if mi < len(cell_texts) and pi < len(cell_texts):
                    mp = cell_texts[mi].strip()
                    party = cell_texts[pi].strip()
                    if mp and party:
                        result[mp.lower()] = party
            if not member_indices:
                for i, text in enumerate(cell_texts[:-1]):
                    nxt = cell_texts[i + 1]
                    if text and nxt and len(nxt) < 30 and not nxt.isdigit():
                        result[text.lower()] = nxt

    time.sleep(0.3)
    return result


def get_party_from_const_page(db_const_name: str, mp_name: str) -> str:
    if db_const_name not in _const_party_cache:
        _const_party_cache[db_const_name] = _fetch_const_page_parties(db_const_name)
    parties = _const_party_cache[db_const_name]
    if not parties:
        return ''
    mp_lower = mp_name.lower()
    if mp_lower in parties:
        return parties[mp_lower]
    mp_last = mp_lower.split()[-1] if mp_lower.split() else ''
    for key, party in parties.items():
        if mp_last and key.split()[-1] == mp_last:
            return party
    return ''


# ── Table parsing ─────────────────────────────────────────────────────────────

def _expand_row(tr, pending):
    """
    Return effective column values for a table row, merging in any cells
    carried forward via rowspan from earlier rows.
    pending is mutated: {col_idx: [text, rows_remaining]}.
    """
    effective = []
    cell_iter = iter(tr.find_all(['td', 'th']))
    col = 0
    while col < 20:
        if col in pending:
            entry = pending[col]
            effective.append(entry[0])
            entry[1] -= 1
            if entry[1] == 0:
                del pending[col]
            col += 1
            continue
        try:
            cell = next(cell_iter)
        except StopIteration:
            break
        text = _FOOTNOTE.sub('', cell.get_text(' ', strip=True))
        rowspan = int(cell.get('rowspan', 1))
        colspan = int(cell.get('colspan', 1))
        for _ in range(colspan):
            effective.append(text)
            if rowspan > 1:
                pending[col] = [text, rowspan - 1]
            col += 1
    return effective


def parse_mp_rows(soup):
    """
    Parse all wikitables on a "List of MPs elected" page.
    Handles rowspan="2" used for two-seat constituencies.

    Returns:
        rows        — list of (const_name_clean, mp_name, party_name)
        seat_counts — Counter of Wikipedia rows per constituency name
    """
    rows = []
    for table in soup.find_all('table', class_='wikitable'):
        pending = {}
        for tr in table.find_all('tr'):
            if not tr.find('td'):
                _expand_row(tr, pending)
                continue

            effective = _expand_row(tr, pending)
            if len(effective) < 2:
                continue

            const_raw  = effective[0]
            mp_text    = _clean(effective[1])
            party_text = _clean(effective[2]) if len(effective) > 2 else ''

            if not mp_text or mp_text.lower() in ('mp', 'member'):
                continue
            if re.match(r'no\s+return|seat\s+vacant|not\s+returned|vacant', mp_text, re.I):
                continue

            const_clean = _clean(const_raw)
            if not const_clean or len(const_clean) <= 1:
                continue

            rows.append((const_clean, mp_text, party_text))

    seat_counts = Counter(r[0] for r in rows)
    return rows, seat_counts


# ── Spreadsheet generation ────────────────────────────────────────────────────

# Elections are split into two groups for separate tabs.
# 1826 is placed in the recent group because it shares the post-Grampound
# constituency list with 1830/1831.
_RECENT_YEARS  = ['1826', '1830', '1831']
_EARLIER_YEARS = ['1802', '1806', '1807', '1812', '1818', '1820']


def generate_spreadsheet(all_scraped: dict, output_path: str, stdout,
                          aliases: dict | None = None,
                          db_names: set | None = None) -> None:
    """
    Write a two-tab Excel workbook from all_scraped.

    all_scraped : {year_str: {const_name: [(mp_name, party_str), ...]}}
    aliases     : CONSTITUENCY_ALIASES dict — pre-populates the Alt Name column
    db_names    : set of all DB Constituency.name values — used to flag unmatched names

    Columns: Constituency | Alt Name | Seat | <year> MP | <year> Party | ...

    Alt Name is pre-populated from aliases where an entry exists, blank otherwise.
    Constituency names not found in db_names (and not in aliases) are highlighted
    in orange so they stand out as needing a mapping entry.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        stdout.write('openpyxl not installed — skipping spreadsheet generation.')
        return

    aliases  = aliases  or {}
    db_names = db_names or set()

    HEADER_FONT   = Font(bold=True)
    MISSING_FILL  = PatternFill(fill_type='solid', fgColor='FFFF99')  # yellow: missing party
    NOMATCH_FILL  = PatternFill(fill_type='solid', fgColor='FFD580')  # orange: no DB match
    GREY_FONT     = Font(color='999999')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tab_groups = [
        ('1826-1831', _RECENT_YEARS),
        ('Pre-1826',  _EARLIER_YEARS),
    ]

    for tab_name, candidate_years in tab_groups:
        years = [y for y in candidate_years if y in all_scraped]
        if not years:
            continue

        ws = wb.create_sheet(tab_name)

        all_consts = sorted({const for y in years for const in all_scraped[y]})

        def max_seats_for(const):
            return max((len(all_scraped[y].get(const, [])) for y in years), default=1)

        # ── Header ────────────────────────────────────────────────────────────
        # Col layout: A=Constituency  B=Alt Name  C=Seat  D=<y0>MP  E=<y0>Party  ...
        header = ['Constituency', 'Alt Name', 'Seat']
        for y in years:
            header += [f'{y} MP', f'{y} Party']
        ws.append(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT
        ws.freeze_panes = 'D2'  # freeze after Alt Name and Seat

        # ── Data rows ─────────────────────────────────────────────────────────
        for const in all_consts:
            n = max_seats_for(const)
            # Determine DB match status for this constituency
            db_name = aliases.get(const, '')
            is_matched = bool(db_name) or (const in db_names)

            for seat_i in range(n):
                row_vals = [
                    const if seat_i == 0 else '',   # A: Constituency
                    db_name if seat_i == 0 else '',  # B: Alt Name
                    seat_i + 1,                      # C: Seat
                ]
                for y in years:
                    mps = all_scraped[y].get(const, [])
                    if seat_i < len(mps):
                        mp_name, party = mps[seat_i]
                        row_vals += [mp_name, party]
                    else:
                        row_vals += ['', '']
                ws.append(row_vals)

                row_idx = ws.max_row

                # Grey constituency name on continuation rows
                if seat_i > 0:
                    ws.cell(row=row_idx, column=1).font = GREY_FONT

                # Orange on constituency + alt name cells when no DB match found
                if seat_i == 0 and not is_matched:
                    ws.cell(row=row_idx, column=1).fill = NOMATCH_FILL
                    ws.cell(row=row_idx, column=2).fill = NOMATCH_FILL

                # Yellow on missing party cells (col 5, 7, 9, ... = 5 + i*2)
                for i in range(len(years)):
                    party_col = 5 + i * 2
                    cell = ws.cell(row=row_idx, column=party_col)
                    if not cell.value:
                        cell.fill = MISSING_FILL

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 5
        for i in range(len(years)):
            ws.column_dimensions[get_column_letter(4 + i * 2)].width = 32
            ws.column_dimensions[get_column_letter(5 + i * 2)].width = 16

    wb.save(output_path)
    stdout.write(f'\nSpreadsheet saved: {output_path}')


# ── Management command ────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = 'Import elected MPs from Wikipedia "List of MPs elected" pages'

    def add_arguments(self, parser):
        group = parser.add_mutually_exclusive_group(required=True)
        group.add_argument('--year', type=str, help='Process a single election year')
        group.add_argument('--all',  action='store_true',
                           help='Process all pre-1832 general elections in the DB')
        parser.add_argument('--dry-run',   action='store_true',
                            help='Show what would be imported without writing to the DB')
        parser.add_argument('--overwrite', action='store_true',
                            help='Delete existing CandidateResult rows before importing')
        parser.add_argument('--fetch-parties', action='store_true',
                            help='When party is missing, fetch the individual constituency '
                                 'Wikipedia page to look it up (slow)')
        parser.add_argument('--spreadsheet', metavar='PATH',
                            help='Save a review spreadsheet to PATH (e.g. review.xlsx). '
                                 'Shows raw scraped data; combine with --dry-run to skip DB import.')

    def handle(self, *args, **options):
        dry_run       = options['dry_run']
        overwrite     = options['overwrite']
        fetch_parties = options['fetch_parties']
        spreadsheet   = options['spreadsheet']

        if options['year']:
            elections = list(Election.objects.filter(year=options['year'], type='GE').order_by('date'))
        else:
            elections = [e for e in Election.objects.filter(type='GE').order_by('date')
                         if e.year and e.year < '1832']

        if not elections:
            self.stderr.write('No matching elections found in the database.')
            return

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — nothing will be written to the DB.\n'))
        if fetch_parties:
            self.stdout.write('Party fallback enabled: will fetch constituency pages for missing parties.\n')

        total_created = 0
        total_skipped = 0
        all_scraped: dict[str, dict[str, list]] = {}
        all_db_names: set[str] = set()  # union of active DB constituency names

        for election in elections:
            self.stdout.write(self.style.MIGRATE_HEADING(f'\n=== {election.year} ==='))

            url, soup = find_wiki_page(election.year)
            if soup is None:
                self.stdout.write(self.style.ERROR(
                    f'  No Wikipedia page found (tried {len(URL_TEMPLATES)} patterns)'))
                continue

            self.stdout.write(f'  {url}')
            rows, wiki_seat_counts = parse_mp_rows(soup)
            self.stdout.write(f'  {len(rows)} MP rows parsed')

            # Collect raw scraped data for spreadsheet (before any DB matching)
            year_data: dict[str, list] = {}
            for const_wiki, mp_name, party_wiki in rows:
                year_data.setdefault(const_wiki, []).append((mp_name, party_wiki))
            all_scraped[election.year] = year_data

            # Constituencies active on the election date
            active_qs = (Constituency.objects
                         .filter(start_date__lte=election.date)
                         .filter(Q(end_date__isnull=True) | Q(end_date__gt=election.date)))
            by_name = {c.name: c for c in active_qs}
            all_db_names.update(by_name.keys())

            party_map = {p.name.lower(): p for p in Party.objects.all()}

            if overwrite and not dry_run:
                deleted, _ = CandidateResult.objects.filter(election=election).delete()
                self.stdout.write(f'  Deleted {deleted} existing rows')

            unmatched_consts  = {}
            unmatched_parties = {}
            seat_mismatches   = []
            party_from_page   = 0
            created = 0
            skipped = 0

            for const_wiki, mp_name, party_wiki in rows:
                # Resolve constituency
                db_const_name = CONSTITUENCY_ALIASES.get(const_wiki, const_wiki)
                const_obj = by_name.get(db_const_name)
                if const_obj is None:
                    unmatched_consts[const_wiki] = None
                    skipped += 1
                    continue

                # Seat count check
                wiki_seats = wiki_seat_counts[const_wiki]
                if wiki_seats != const_obj.seats:
                    seat_mismatches.append(
                        f'{const_wiki}: Wikipedia={wiki_seats}, DB={const_obj.seats}'
                    )

                # Resolve party (optional constituency-page fallback)
                if not party_wiki and fetch_parties:
                    self.stdout.write(f'  Fetching party for {mp_name} ({db_const_name})...',
                                      ending='\r')
                    party_wiki = get_party_from_const_page(db_const_name, mp_name)
                    if party_wiki:
                        party_from_page += 1

                party_key = PARTY_ALIASES.get(party_wiki.lower(), party_wiki.lower())
                party_obj = party_map.get(party_key)
                if party_obj is None:
                    unmatched_parties[party_wiki or '(empty)'] = None
                    skipped += 1
                    continue

                if not dry_run:
                    CandidateResult.objects.get_or_create(
                        election=election,
                        constituency=const_obj,
                        candidate=mp_name,
                        defaults={'party': party_obj, 'elected': True},
                    )
                created += 1

            total_created += created
            total_skipped += skipped

            self.stdout.write(
                f'  {"Would create" if dry_run else "Created"}: {created}  '
                f'Skipped (no match): {skipped}'
                + (f'  Party from constituency page: {party_from_page}' if party_from_page else '')
            )

            if seat_mismatches:
                self.stdout.write(self.style.WARNING(
                    f'  Seat count mismatches ({len(seat_mismatches)}):'))
                for msg in seat_mismatches:
                    self.stdout.write(f'    {msg}')

            if unmatched_consts:
                self.stdout.write(self.style.WARNING(
                    f'  Unmatched constituencies ({len(unmatched_consts)}) '
                    '— add to CONSTITUENCY_ALIASES:'))
                for name in sorted(unmatched_consts):
                    close = [n for n in by_name if n.lower().startswith(name[:4].lower())][:3]
                    hint = f'  (similar in DB: {close})' if close else ''
                    self.stdout.write(f'    {name!r}{hint}')

            if unmatched_parties:
                self.stdout.write(self.style.WARNING(
                    f'  Unmatched parties ({len(unmatched_parties)}) '
                    '— add to PARTY_ALIASES or create in DB:'))
                for name in sorted(unmatched_parties):
                    self.stdout.write(f'    {name!r}')

            time.sleep(0.5)

        if spreadsheet:
            generate_spreadsheet(all_scraped, spreadsheet, self.stdout,
                                 aliases=CONSTITUENCY_ALIASES,
                                 db_names=all_db_names)

        self.stdout.write(self.style.SUCCESS(
            f'\nDone.  Total {"would create" if dry_run else "created"}: {total_created}  '
            f'Skipped: {total_skipped}'
        ))
